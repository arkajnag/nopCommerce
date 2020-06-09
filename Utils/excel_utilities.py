import openpyxl


def getRowCount(filePath, sheetName):
    wb = openpyxl.load_workbook(filename=filePath)
    sheet = wb[sheetName]
    return sheet.max_row


def getColumnCount(filePath, sheetName):
    wb = openpyxl.load_workbook(filename=filePath)
    sheet = wb[sheetName]
    return sheet.max_column


def readDatafromExcel(filePath, sheetName, xRow, yCol):
    wb = openpyxl.load_workbook(filename=filePath)
    sheet = wb[sheetName]
    return sheet.cell(row=xRow, column=yCol).value

def writeDataIntoExcel(filePath, sheetName, xRow, yCol, data):
    wb = openpyxl.load_workbook(filename=filePath)
    sheet = wb[sheetName]
    sheet.cell(row=xRow, column=yCol).value = data
    wb.save(filePath)
